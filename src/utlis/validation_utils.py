import pandas as pd
import openpyxl
import re
from typing import List, Dict, Any, Union

def validate_kpi_data(extracted_kpis: List[Dict[str, Any]], file_path_to_validate_against: str, target_kpis: List[str]) -> Dict[str, Any]:
    """
    Validates extracted KPI data against the original Excel file.
    Uses LLM-provided row and column numbers for cross-validation.

    Args:
        extracted_kpis (List[Dict[str, Any]]): List of KPIs extracted by the LLM, including row_number and column_number.
        excel_file_path (str): Path to the original Excel file.
        target_kpis (List[str]): The list of KPIs the agent was asked to extract.

    Returns:
        Dict[str, Any]: A dictionary containing validated KPIs and unextracted numbers.
    """
    print(f"DEBUG: Starting KPI validation for file: {file_path_to_validate_against}")
    
    validated_kpis = []
    unextracted_numbers = []

    # --- Tier 1: Structural Validation ---
    structurally_valid_kpis = []
    print(f"DEBUG: Starting structural validation for {len(extracted_kpis)} extracted KPIs.")
    for i, kpi_data in enumerate(extracted_kpis):
        print(f"DEBUG: Processing KPI {i+1}/{len(extracted_kpis)}: {kpi_data.get('kpi')} - Value: {kpi_data.get('value')}")
        kpi_name = kpi_data.get("kpi")
        kpi_value = kpi_data.get("value")
        kpi_period = kpi_data.get("period")
        source_file = kpi_data.get("source_file")
        row_number = kpi_data.get("row_number")
        column_number = kpi_data.get("column_number")

        validation_status = "Valid"
        notes = []

        if not kpi_name or not isinstance(kpi_name, str) or not kpi_name.strip():
            validation_status = "INVALID_STRUCTURE"
            notes.append("KPI name is missing or invalid.")
        
        if kpi_value is None:
            validation_status = "INVALID_STRUCTURE"
            notes.append("Value is missing.")
        elif isinstance(kpi_value, str):
            try:
                float(kpi_value.replace(",", "")) # Attempt to convert to float, ignoring commas
            except ValueError:
                validation_status = "INVALID_STRUCTURE"
                notes.append("Value is not a valid number format.")
        elif not isinstance(kpi_value, (int, float)):
            validation_status = "INVALID_STRUCTURE"
            notes.append("Value is not a valid number format.")
        
        if not kpi_period or not isinstance(kpi_period, str) or not kpi_period.strip():
            validation_status = "INVALID_STRUCTURE"
            notes.append("Period is missing or invalid.")
        
        if row_number is None or not isinstance(row_number, int) or row_number <= 0:
            validation_status = "INVALID_STRUCTURE"
            notes.append("Row number is missing or invalid.")

        if column_number is None or not isinstance(column_number, int) or column_number <= 0:
            validation_status = "INVALID_STRUCTURE"
            notes.append("Column number is missing or invalid.")

        if validation_status == "INVALID_STRUCTURE":
            print(f"DEBUG: KPI '{kpi_name}' failed structural validation. Notes: {notes}")
            validated_kpis.append({
                "kpi": kpi_name,
                "value": kpi_value,
                "period": kpi_period,
                "source_file": source_file,
                "row_number": row_number,
                "column_number": column_number,
                "validation_status": validation_status,
                "notes": notes,
                "validated": False # Set to False for structural invalidation
            })
        else:
            print(f"DEBUG: KPI '{kpi_name}' passed structural validation.")
            # Initialize 'validated' to True for structurally valid KPIs, will be updated during cross-validation
            kpi_data["validation_status"] = validation_status # Add validation_status
            kpi_data["notes"] = notes # Add notes
            kpi_data["validated"] = True
            structurally_valid_kpis.append(kpi_data)

    print(f"DEBUG: Structural validation completed. Found {len(structurally_valid_kpis)} structurally valid KPIs.")
    if not structurally_valid_kpis:
        print("DEBUG: No structurally valid KPIs to cross-validate.")
        return {
            "validated_kpis": validated_kpis,
            "unextracted_numbers": []
        }

    # Determine if the file is an Excel or CSV and load accordingly
    is_excel = file_path_to_validate_against.endswith((".xlsx", ".xls"))
    
    if is_excel:
        try:
            workbook = openpyxl.load_workbook(file_path_to_validate_against, data_only=True)
            sheets_to_process = workbook.sheetnames
            print(f"DEBUG: Loaded Excel workbook {file_path_to_validate_against}. Sheets: {sheets_to_process}")
        except Exception as e:
            print(f"ERROR: Could not load Excel workbook {file_path_to_validate_against}: {e}")
            for kpi_data in structurally_valid_kpis:
                kpi_data["validation_status"] = "FILE_LOAD_ERROR"
                kpi_data["notes"].append(f"Could not load Excel file: {e}")
                kpi_data["validated"] = False
                validated_kpis.append(kpi_data)
            return {
                "validated_kpis": validated_kpis,
                "unextracted_numbers": []
            }
    else: # Assume CSV
        try:
            df = pd.read_csv(file_path_to_validate_against)
            # For CSV, we treat the entire file as one "sheet" for validation purposes
            # We'll create a dummy workbook-like structure or adapt the logic
            workbook = None # Indicate no openpyxl workbook
            sheets_to_process = [file_path_to_validate_against] # Use the file path as a dummy sheet name
            print(f"DEBUG: Loaded CSV file {file_path_to_validate_against}.")
        except Exception as e:
            print(f"ERROR: Could not load CSV file {file_path_to_validate_against}: {e}")
            for kpi_data in structurally_valid_kpis:
                kpi_data["validation_status"] = "FILE_LOAD_ERROR"
                kpi_data["notes"].append(f"Could not load CSV file: {e}")
                kpi_data["validated"] = False
                validated_kpis.append(kpi_data)
            return {
                "validated_kpis": validated_kpis,
                "unextracted_numbers": []
            }

    potential_unextracted_kpi_values = set()

    print(f"DEBUG: Iterating through sheets to find potential unextracted KPI values.")
    for sheet_identifier in sheets_to_process:
        print(f"DEBUG: Processing sheet/file: {sheet_identifier}")
        
        if is_excel:
            sheet = workbook[sheet_identifier]
            # Logic for Excel remains largely the same
            for target_kpi_name in target_kpis:
                print(f"DEBUG: Searching for target KPI '{target_kpi_name}' in sheet '{sheet_identifier}'.")
                kpi_row_index = -1
                for r_idx, row in enumerate(sheet.iter_rows()):
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and target_kpi_name.lower() == str(cell.value).lower():
                            kpi_row_index = r_idx
                            print(f"DEBUG: Found target KPI '{target_kpi_name}' at row {r_idx + 1}.")
                            break
                    if kpi_row_index != -1:
                        break
                
                if kpi_row_index != -1:
                    print(f"DEBUG: Extracting potential values from row {kpi_row_index + 2} for KPI '{target_kpi_name}'.")
                    for c_idx, cell in enumerate(sheet[kpi_row_index + 1]):
                        if isinstance(cell.value, (int, float)):
                            potential_unextracted_kpi_values.add(str(cell.value).replace(",", ""))
                            print(f"DEBUG: Added numeric value '{cell.value}' to potential unextracted values.")
                        elif isinstance(cell.value, str):
                            num_str_cleaned = cell.value.replace(",", "")
                            if re.match(r'^\d+(\.\d+)?$', num_str_cleaned):
                                potential_unextracted_kpi_values.add(num_str_cleaned)
                                print(f"DEBUG: Added string numeric value '{cell.value}' to potential unextracted values.")
        else: # CSV processing
            # For CSV, the entire DataFrame is the "sheet"
            # We need to iterate through rows and columns of the DataFrame
            for target_kpi_name in target_kpis:
                print(f"DEBUG: Searching for target KPI '{target_kpi_name}' in CSV: {sheet_identifier}.")
                kpi_row_index = -1
                # Iterate through DataFrame rows to find KPI name
                for r_idx, row_series in df.iterrows():
                    for c_idx, cell_value in row_series.items():
                        if cell_value and isinstance(cell_value, str) and target_kpi_name.lower() == str(cell_value).lower():
                            kpi_row_index = r_idx
                            print(f"DEBUG: Found target KPI '{target_kpi_name}' at row {r_idx + 1}.")
                            break
                    if kpi_row_index != -1:
                        break
                
                if kpi_row_index != -1:
                    print(f"DEBUG: Extracting potential values from row {kpi_row_index + 1} for KPI '{target_kpi_name}' in CSV.")
                    # Iterate through cells in the row after the KPI name
                    for c_idx, cell_value in df.iloc[kpi_row_index].items():
                        if isinstance(cell_value, (int, float)):
                            potential_unextracted_kpi_values.add(str(cell_value).replace(",", ""))
                            print(f"DEBUG: Added numeric value '{cell_value}' to potential unextracted values.")
                        elif isinstance(cell_value, str):
                            num_str_cleaned = cell_value.replace(",", "")
                            if re.match(r'^\d+(\.\d+)?$', num_str_cleaned):
                                potential_unextracted_kpi_values.add(num_str_cleaned)
                                print(f"DEBUG: Added string numeric value '{cell_value}' to potential unextracted values.")

    print(f"DEBUG: Starting cross-validation for {len(structurally_valid_kpis)} structurally valid KPIs.")
    for kpi_data in structurally_valid_kpis:
        if kpi_data.get("validation_status") == "Valid": # Ensure it's still valid from previous checks
            kpi_name = kpi_data["kpi"]
            kpi_value = kpi_data["value"]
            kpi_period = kpi_data["period"]
            row_number = kpi_data["row_number"]
            column_number = kpi_data["column_number"]
            extracted_value_str = str(kpi_value).replace(",", "")
            print(f"DEBUG: Cross-validating KPI '{kpi_name}' (Value: '{extracted_value_str}') at R{row_number}C{column_number}.")

            try:
                # Access the sheet using the name from the extracted data's source_file
                # This assumes sheet_name can be derived from source_file or is consistent
                # For now, we'll iterate through all sheets to find the cell
                source_value_found = False
                for sheet_identifier_for_cell_access in sheets_to_process:
                    if is_excel:
                        sheet = workbook[sheet_identifier_for_cell_access]
                        try:
                            source_value_cell = sheet.cell(row=row_number, column=column_number).value
                            if source_value_cell is not None:
                                source_value_str_from_source = str(source_value_cell).replace(",", "")
                                
                                # Compare values
                                if extracted_value_str == source_value_str_from_source:
                                    kpi_data["validation_status"] = "Valid"
                                    kpi_data["validated"] = True
                                    print(f"DEBUG: KPI '{kpi_name}' validation result: True (Exact match)")
                                else:
                                    # Attempt numeric comparison if both are numeric
                                    try:
                                        if abs(float(extracted_value_str) - float(source_value_str_from_source)) < 1e-9:
                                            kpi_data["validation_status"] = "Valid"
                                            kpi_data["validated"] = True
                                            print(f"DEBUG: KPI '{kpi_name}' validation result: True (Numeric match)")
                                        else:
                                            kpi_data["validation_status"] = "VALUE_MISMATCH"
                                            kpi_data["notes"].append(f"Value mismatch. Expected: '{source_value_str_from_source}', Got: '{extracted_value_str}' at R{row_number}C{column_number}.")
                                            kpi_data["validated"] = False
                                            print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value mismatch)")
                                    except ValueError:
                                        # If not numeric, or conversion fails, compare as strings
                                        kpi_data["validation_status"] = "VALUE_MISMATCH"
                                        kpi_data["notes"].append(f"Value mismatch. Expected: '{source_value_str_from_source}', Got: '{extracted_value_str}' at R{row_number}C{column_number}.")
                                        kpi_data["validated"] = False
                                        print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value mismatch, non-numeric)")
                                source_value_found = True
                                break # Value found in this sheet, no need to check others
                            else:
                                kpi_data["validation_status"] = "VALUE_MISSING_IN_SOURCE"
                                kpi_data["notes"].append(f"Value cell is empty in source Excel at R{row_number}C{column_number}.")
                                kpi_data["validated"] = False
                                source_value_found = True
                                print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value missing in source)")
                                break
                        except Exception as e:
                            print(f"WARNING: Error accessing cell R{row_number}C{column_number} in sheet {sheet_identifier_for_cell_access}: {e}")
                            pass
                    else: # CSV validation
                        try:
                            # For CSV, row_number and column_number are 1-based from LLM, pandas is 0-based
                            source_value_cell = df.iloc[row_number - 1, column_number - 1]
                            if source_value_cell is not None:
                                source_value_str_from_source = str(source_value_cell).replace(",", "")
                                
                                # Compare values
                                if extracted_value_str == source_value_str_from_source:
                                    kpi_data["validation_status"] = "Valid"
                                    kpi_data["validated"] = True
                                    print(f"DEBUG: KPI '{kpi_name}' validation result: True (Exact match)")
                                else:
                                    try:
                                        if abs(float(extracted_value_str) - float(source_value_str_from_source)) < 1e-9:
                                            kpi_data["validation_status"] = "Valid"
                                            kpi_data["validated"] = True
                                            print(f"DEBUG: KPI '{kpi_name}' validation result: True (Numeric match)")
                                        else:
                                            kpi_data["validation_status"] = "VALUE_MISMATCH"
                                            kpi_data["notes"].append(f"Value mismatch. Expected: '{source_value_str_from_source}', Got: '{extracted_value_str}' at R{row_number}C{column_number}.")
                                            kpi_data["validated"] = False
                                            print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value mismatch)")
                                    except ValueError:
                                        kpi_data["validation_status"] = "VALUE_MISMATCH"
                                        kpi_data["notes"].append(f"Value mismatch. Expected: '{source_value_str_from_source}', Got: '{extracted_value_str}' at R{row_number}C{column_number}.")
                                        kpi_data["validated"] = False
                                        print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value mismatch, non-numeric)")
                                source_value_found = True
                                break
                            else:
                                kpi_data["validation_status"] = "VALUE_MISSING_IN_SOURCE"
                                kpi_data["notes"].append(f"Value cell is empty in source CSV at R{row_number}C{column_number}.")
                                kpi_data["validated"] = False
                                source_value_found = True
                                print(f"DEBUG: KPI '{kpi_name}' validation result: False (Value missing in source CSV)")
                                break
                        except IndexError:
                            print(f"WARNING: Cell R{row_number}C{column_number} out of bounds for CSV file {sheet_identifier_for_cell_access}.")
                            # This means the row/column from LLM is invalid for this CSV
                            pass
                        except Exception as e:
                            print(f"WARNING: Error accessing cell R{row_number}C{column_number} in CSV {sheet_identifier_for_cell_access}: {e}")
                            pass
                
                if not source_value_found:
                    print(f"DEBUG: Value not found at R{row_number}C{column_number} in any sheet/file for KPI '{kpi_name}'.")
                    kpi_data["validation_status"] = "LOCATION_NOT_FOUND_IN_SOURCE"
                    kpi_data["notes"].append(f"Could not find value at R{row_number}C{column_number} in any sheet/file of the source.")
                    kpi_data["validated"] = False
                    print(f"DEBUG: KPI '{kpi_name}' validation result: False (Location not found)")

            except Exception as e:
                print(f"ERROR: Unexpected error during cross-validation for KPI '{kpi_name}': {e}")
                kpi_data["validation_status"] = "CROSS_VALIDATION_ERROR"
                kpi_data["notes"].append(f"Error during cross-validation for KPI '{kpi_name}': {e}")
                kpi_data["validated"] = False
            
            validated_kpis.append(kpi_data)

    print(f"DEBUG: Cross-validation completed. Total validated KPIs: {len(validated_kpis)}.")

    extracted_values_from_current_file = {
        str(kpi_data.get("value")).replace(",", "")
        for kpi_data in extracted_kpis
        if kpi_data.get("value") is not None
    }

    print(f"DEBUG: Identifying unextracted numbers from potential values.")
    for num_str in potential_unextracted_kpi_values:
        if num_str not in extracted_values_from_current_file:
            unextracted_numbers.append(num_str)
            print(f"DEBUG: Identified unextracted number: {num_str}")

    print(f"DEBUG: KPI validation completed for file: {file_path_to_validate_against}. Found {len(unextracted_numbers)} unextracted numbers.")
    return {
        "validated_kpis": validated_kpis,
        "unextracted_numbers": list(unextracted_numbers)
    }
