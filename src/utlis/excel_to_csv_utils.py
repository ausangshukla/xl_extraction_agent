import openpyxl
import pandas as pd
import os
from openpyxl.utils.cell import column_index_from_string, get_column_letter

def convert_excel_to_csv(excel_file_path, output_folder):
    """
    Converts a single Excel file (all its worksheets) into separate CSV files.

    Args:
        excel_file_path (str): The path to the input .xlsx or .xls file.
        output_folder (str): The path to the folder where the CSV files will be saved.
    """
    if not os.path.exists(excel_file_path):
        print(f"Error: Excel file '{excel_file_path}' not found.")
        return

    if not os.path.isdir(output_folder):
        os.makedirs(output_folder, exist_ok=True)

    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        base_name = os.path.splitext(os.path.basename(excel_file_path))[0]

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Determine visible rows and columns
            visible_rows = []
            for row_idx in range(1, sheet.max_row + 1):
                if not sheet.row_dimensions[row_idx].hidden:
                    visible_rows.append(row_idx)

            visible_col_letters = []
            for col_idx in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if not sheet.column_dimensions[col_letter].hidden:
                    visible_col_letters.append(col_letter)

            # Read the Excel sheet into a pandas DataFrame, skipping hidden rows and columns
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None)

            # Filter DataFrame to include only visible rows and columns
            df_filtered = df.iloc[[r - 1 for r in visible_rows if (r - 1) < len(df)]]
            
            # Filter columns by letter
            df_columns_map = {get_column_letter(i+1): i for i in range(df_filtered.shape[1])}
            cols_to_keep_indices = [df_columns_map[col_letter] for col_letter in visible_col_letters if col_letter in df_columns_map]
            df_filtered = df_filtered.iloc[:, cols_to_keep_indices]

            # Remove rows that are entirely blank (contain only NaN values)
            df_filtered.dropna(how='all', inplace=True)
            # Remove columns that are entirely blank (contain only NaN values)
            df_filtered.dropna(axis=1, how='all', inplace=True)

            # Promote the first non-blank row to header if it exists
            if not df_filtered.empty:
                df_filtered.columns = df_filtered.iloc[0]
                df_filtered = df_filtered[1:].reset_index(drop=True)

            # Construct the CSV filename using the base name of the Excel file and the sheet name
            csv_filename = f"{base_name}_{sheet_name}.csv"
            # Create the full path for the output CSV file
            csv_file_path = os.path.join(output_folder, csv_filename)
            # Save the DataFrame to a CSV file, without writing the DataFrame index
            df_filtered.to_csv(csv_file_path, index=False)

            print(f"Converted '{sheet_name}' from '{os.path.basename(excel_file_path)}' to '{csv_filename}'")
    except Exception as e:
        print(f"Error processing '{os.path.basename(excel_file_path)}': {e}")

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Convert Excel files to CSVs.")
    parser.add_argument("excel_file_path", type=str, help="Path to the Excel file (.xlsx or .xls).")
    parser.add_argument("output_folder", type=str, help="Path to the output folder for CSVs.")
    args = parser.parse_args()

    convert_excel_to_csv(args.excel_file_path, args.output_folder)